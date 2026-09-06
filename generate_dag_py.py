#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dag_generation_input.xlsx를 읽어서 Airflow + dbt(Cosmos) DAG Python 코드를 생성한다.

입력:
  - dag_generation_input.xlsx

출력:
  - DAG 1개당 Python 파일 1개

기본 동작:
  - Excel의 4개 시트(DAG_Mapping, Sensor_Notify, Validation_Policy, Review_Log)를 읽는다.
  - seq_name 기준으로 한 줄을 합쳐 DAG 스펙을 만든다.
  - 규칙표에 따라 DAG 코드를 렌더링한다.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl이 필요합니다: pip install openpyxl") from exc


SHEETS = ["DAG_Mapping", "Sensor_Notify", "Validation_Policy", "Review_Log"]


@dataclass
class DagSpec:
    seq_name: str
    business_name: str
    dag_id: str
    dag_name: str
    schedule_type: str
    schedule_interval: str
    priority: str
    dbt_group: str
    dbt_selector: str
    model_mapping: str
    start_mode: str
    sensor_type: str
    sensor_target: str
    external_dependency: str
    notify_on_success: str
    notify_on_failure: str
    notify_channel: str
    notify_recipients: str
    failure_policy: str
    validation_mode: str
    validation_timing: str
    comparison_reference: str
    retry_policy: str
    exception_rule: str
    verification_comment: str
    review_status: str
    review_comment: str
    owner: str
    reviewed_at: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Y" if value else "N"
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def slugify(value: str) -> str:
    value = normalize_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value


def safe_filename(value: str) -> str:
    value = normalize_text(value)
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", "_", value).strip("._")
    return value or "output"


def split_retry_policy(value: str) -> Tuple[int, int]:
    text = normalize_text(value)
    retries = 1
    delay_min = 5

    m = re.search(r"(\d+)\s*회", text)
    if m:
        retries = int(m.group(1))
    m = re.search(r"(\d+)\s*분", text)
    if m:
        delay_min = int(m.group(1))
    return retries, delay_min


def schedule_from_type(schedule_type: str) -> str:
    mapping = {
        "daily": "0 1 * * *",
        "weekly": "0 1 * * 1",
        "monthly": "0 1 1 * *",
        "manual": None,
        "event": None,
    }
    return mapping.get(normalize_text(schedule_type).lower(), None)


def read_rows(ws) -> List[Dict[str, Any]]:
    headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        data = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if any(v not in (None, "") for v in data.values()):
            rows.append(data)
    return rows


def load_excel(input_path: Path, seq_names: Optional[Sequence[str]]) -> Dict[str, DagSpec]:
    wb = load_workbook(input_path)
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            raise SystemExit(f"필수 시트가 없습니다: {sheet}")

    rows_by_seq: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for sheet in SHEETS:
        for row in read_rows(wb[sheet]):
            seq_name = normalize_text(row.get("seq_name"))
            if not seq_name:
                continue
            if seq_names and seq_name not in seq_names:
                continue
            rows_by_seq.setdefault(seq_name, {})[sheet] = row

    specs: Dict[str, DagSpec] = {}
    for seq_name, bundle in rows_by_seq.items():
        if not all(sheet in bundle for sheet in SHEETS):
            missing = [s for s in SHEETS if s not in bundle]
            raise SystemExit(f"{seq_name}: 필요한 시트 행이 부족합니다: {', '.join(missing)}")
        specs[seq_name] = build_spec(bundle)
    return specs


def build_spec(bundle: Dict[str, Dict[str, Any]]) -> DagSpec:
    dag = bundle["DAG_Mapping"]
    sensor = bundle["Sensor_Notify"]
    validation = bundle["Validation_Policy"]
    review = bundle["Review_Log"]

    seq_name = normalize_text(dag.get("seq_name") or sensor.get("seq_name") or validation.get("seq_name") or review.get("seq_name"))
    business_name = normalize_text(dag.get("business_name"))
    dag_id = normalize_text(dag.get("dag_id")) or slugify(seq_name)
    dag_name = normalize_text(dag.get("dag_name")) or f"{seq_name} - {business_name}".strip(" -")

    schedule_type = normalize_text(dag.get("schedule_type")).lower()
    schedule_interval = normalize_text(dag.get("schedule_interval")) or (schedule_from_type(schedule_type) or "")
    priority = normalize_text(dag.get("priority")) or "medium"
    dbt_group = normalize_text(dag.get("dbt_group")) or slugify(seq_name)
    dbt_selector = normalize_text(dag.get("dbt_selector")) or (f"tag:{dbt_group}" if dbt_group else "")
    model_mapping = normalize_text(dag.get("model_mapping"))

    start_mode = normalize_text(sensor.get("start_mode")).lower()
    sensor_type = normalize_text(sensor.get("sensor_type"))
    sensor_target = normalize_text(sensor.get("sensor_target"))
    external_dependency = normalize_text(sensor.get("external_dependency"))
    notify_on_success = normalize_text(sensor.get("notify_on_success")) or "N"
    notify_on_failure = normalize_text(sensor.get("notify_on_failure")) or "N"
    notify_channel = normalize_text(sensor.get("notify_channel"))
    notify_recipients = normalize_text(sensor.get("notify_recipients"))
    failure_policy = normalize_text(sensor.get("failure_policy")) or "stop"

    validation_mode = normalize_text(validation.get("validation_mode")) or "dbt_test"
    validation_timing = normalize_text(validation.get("validation_timing"))
    comparison_reference = normalize_text(validation.get("comparison_reference"))
    retry_policy = normalize_text(validation.get("retry_policy"))
    exception_rule = normalize_text(validation.get("exception_rule"))
    verification_comment = normalize_text(validation.get("verification_comment"))

    review_status = normalize_text(review.get("review_status")) or "needs_review"
    review_comment = normalize_text(review.get("review_comment"))
    owner = normalize_text(review.get("owner")) or normalize_text(dag.get("owner"))
    reviewed_at = normalize_text(review.get("reviewed_at"))

    return DagSpec(
        seq_name=seq_name,
        business_name=business_name,
        dag_id=dag_id,
        dag_name=dag_name,
        schedule_type=schedule_type,
        schedule_interval=schedule_interval,
        priority=priority,
        dbt_group=dbt_group,
        dbt_selector=dbt_selector,
        model_mapping=model_mapping,
        start_mode=start_mode,
        sensor_type=sensor_type,
        sensor_target=sensor_target,
        external_dependency=external_dependency,
        notify_on_success=notify_on_success,
        notify_on_failure=notify_on_failure,
        notify_channel=notify_channel,
        notify_recipients=notify_recipients,
        failure_policy=failure_policy,
        validation_mode=validation_mode,
        validation_timing=validation_timing,
        comparison_reference=comparison_reference,
        retry_policy=retry_policy,
        exception_rule=exception_rule,
        verification_comment=verification_comment,
        review_status=review_status,
        review_comment=review_comment,
        owner=owner,
        reviewed_at=reviewed_at,
    )


def parse_validation_comment(spec: DagSpec) -> str:
    comment = spec.verification_comment
    if comment:
        return comment
    return f"검증 방식={spec.validation_mode}, 비교 기준={spec.comparison_reference}"


def render_dag_code(spec: DagSpec, strict_approved_only: bool = False) -> str:
    schedule = spec.schedule_interval if spec.schedule_interval else None
    retries, delay_min = split_retry_policy(spec.retry_policy)
    owner = spec.owner or "TODO_OWNER"
    project_dir = f'/opt/airflow/dbt/{spec.dbt_group}'
    profile_name = spec.dbt_group or "TODO_PROFILE"
    profile_target = "prod"

    notes: List[str] = []
    if strict_approved_only and spec.review_status != "approved":
        notes.append("# NOTE: review_status != approved, scaffold mode")
    if not spec.schedule_interval:
        notes.append("# NOTE: schedule_interval was inferred from schedule_type")
    if not spec.sensor_type:
        notes.append("# NOTE: sensor_type is missing; using start_mode/manual scaffold")
    if not spec.notify_channel:
        notes.append("# NOTE: notify_channel is missing; notify tasks are placeholders")
    if not spec.model_mapping:
        notes.append("# NOTE: model_mapping is missing; dbt graph dependency is used")

    start_flow = _render_start_flow(spec)
    dbt_render_config = f'RenderConfig(select=[{spec.dbt_selector!r}])' if spec.dbt_selector else 'RenderConfig(select=[])'
    lines: List[str] = []
    lines.append("# Auto-generated from dag_generation_input.xlsx")
    lines.append(f"# seq_name: {spec.seq_name}")
    lines.append(f"# review_status: {spec.review_status}")
    for note in notes:
        lines.append(note)
    lines.append("")
    lines.append("from __future__ import annotations")
    lines.append("")
    lines.append("from datetime import datetime, timedelta")
    lines.append("")
    lines.append("from airflow import DAG")
    lines.append("from airflow.operators.bash import BashOperator")
    lines.append("from airflow.operators.empty import EmptyOperator")
    lines.append("from airflow.operators.python import BranchPythonOperator")
    lines.append("from cosmos import DbtTaskGroup, ExecutionConfig, ProfileConfig, ProjectConfig, RenderConfig")
    lines.append("from cosmos.profiles import PostgresUserPasswordProfileMapping")
    lines.append("")
    lines.append(f"DAG_ID = {spec.dag_id!r}")
    lines.append(f"DAG_NAME = {spec.dag_name!r}")
    lines.append(f"DESCRIPTION = {spec.business_name!r} + ' 업무를 Airflow + dbt(Cosmos)로 오케스트레이션한다.'")
    lines.append("START_DATE = datetime(2026, 9, 5)")
    lines.append(f"REVIEW_STATUS = {spec.review_status!r}")
    lines.append(f"REVIEW_COMMENT = {spec.review_comment!r}")
    lines.append("")
    lines.append("DEFAULT_ARGS = {")
    lines.append(f"    'owner': {owner!r},")
    lines.append("    'depends_on_past': False,")
    lines.append(f"    'retries': {retries},")
    lines.append(f"    'retry_delay': timedelta(minutes={delay_min}),")
    lines.append("    'email_on_failure': False,")
    lines.append("    'email_on_retry': False,")
    lines.append("}")
    lines.append("")
    lines.append("")
    lines.append("def choose_route() -> str:")
    lines.append(f"    return {start_flow!r}")
    lines.append("")
    lines.append("")
    lines.append("with DAG(")
    lines.append("    dag_id=DAG_ID,")
    lines.append("    dag_display_name=DAG_NAME,")
    lines.append("    description=DESCRIPTION,")
    lines.append("    start_date=START_DATE,")
    lines.append(f"    schedule={schedule!r},")
    lines.append("    catchup=False,")
    lines.append("    default_args=DEFAULT_ARGS,")
    lines.append(f"    tags={[spec.schedule_type or 'custom', spec.dbt_group, 'dbt', 'airflow']!r},")
    lines.append("    max_active_runs=1,")
    lines.append(") as dag:")
    lines.append("    start = EmptyOperator(task_id='start')")
    lines.append("")
    lines.append("    route = BranchPythonOperator(")
    lines.append("        task_id='route',")
    lines.append("        python_callable=choose_route,")
    lines.append("    )")
    lines.append("")
    lines.append("    manual_start = EmptyOperator(task_id='manual_start')")
    lines.append("")
    lines.append("    dbt_build = DbtTaskGroup(")
    lines.append(f"        group_id={('dbt_' + spec.dbt_group)[:50]!r},")
    lines.append(f"        project_config=ProjectConfig({project_dir!r}),")
    lines.append("        profile_config=ProfileConfig(")
    lines.append(f"            profile_name={profile_name!r},")
    lines.append(f"            target_name={profile_target!r},")
    lines.append("            profile_mapping=PostgresUserPasswordProfileMapping(")
    lines.append("                conn_id='TODO_WAREHOUSE_CONN_ID',")
    lines.append("                profile_args={")
    lines.append("                    'schema': 'TODO_SCHEMA',")
    lines.append("                    'database': 'TODO_DATABASE',")
    lines.append("                },")
    lines.append("            ),")
    lines.append("        ),")
    lines.append("        execution_config=ExecutionConfig(")
    lines.append("            dbt_executable_path='/usr/local/bin/dbt',")
    lines.append("        ),")
    lines.append(f"        render_config={dbt_render_config},")
    lines.append("    )")
    lines.append("")
    lines.append("    validate = BashOperator(")
    lines.append("        task_id='validate',")
    lines.append(f"        bash_command={_render_validation_cmd(spec)!r},")
    lines.append(f"        cwd={project_dir!r},")
    lines.append("    )")
    lines.append("")
    lines.append("    notify_success = EmptyOperator(task_id='notify_success')")
    lines.append("    notify_failure = EmptyOperator(task_id='notify_failure')")
    lines.append("")
    lines.append("    end = EmptyOperator(task_id='end', trigger_rule='none_failed_min_one_success')")
    lines.append("")
    lines.append("    start >> route")
    lines.append("    route >> manual_start >> dbt_build >> validate >> notify_success >> end")
    lines.append("    validate >> notify_failure >> end")
    lines.append("")
    lines.append("    # review / note")
    lines.append("    # TODO: connect notify_channel / notify_recipients if available")
    lines.append("    # TODO: refine model_mapping when populated")
    lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _render_start_flow(spec: DagSpec) -> str:
    if spec.start_mode == "file_sensor" and spec.sensor_type:
        return "wait_input_file"
    if spec.start_mode == "dependency_sensor" and spec.sensor_type:
        return "wait_dependency"
    if spec.start_mode == "event" and spec.sensor_type:
        return "wait_event"
    return "manual_start"


def _render_validation_cmd(spec: DagSpec) -> str:
    selector = spec.dbt_selector or f"tag:{spec.dbt_group}"
    if spec.validation_mode == "dbt_test":
        return f"cd /opt/airflow/dbt/{spec.dbt_group} && dbt test --select {selector}"
    if spec.validation_mode == "count":
        return f"echo 'TODO count validation for {spec.seq_name}'"
    if spec.validation_mode == "sample":
        return f"echo 'TODO sample validation for {spec.seq_name}'"
    if spec.validation_mode == "report":
        return f"echo 'TODO report validation for {spec.seq_name}'"
    return f"echo 'TODO validation for {spec.seq_name}'"


def write_dag_file(output_path: Path, code: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(code, encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="dag_generation_input.xlsx를 Airflow DAG Python 코드로 변환한다")
    p.add_argument("--input", default="dag_generation_input.xlsx", help="입력 Excel 경로")
    p.add_argument("--output-dir", default="dags", help="출력 디렉터리")
    p.add_argument("--output", help="단일 파일 출력 경로")
    p.add_argument("--seq-name", action="append", dest="seq_names", help="대상 seq_name (여러 번 지정 가능)")
    p.add_argument("--split-by-seq", action="store_true", help="seq_name별로 개별 파일 생성")
    p.add_argument("--dry-run", action="store_true", help="파일 저장 없이 코드만 출력")
    p.add_argument("--overwrite", action="store_true", help="기존 출력 파일 덮어쓰기")
    p.add_argument("--strict-approved-only", action="store_true", help="approved 행만 최종본 생성")
    return p


def generate_one(spec: DagSpec, output_path: Path, strict_approved_only: bool, dry_run: bool, overwrite: bool) -> Path:
    code = render_dag_code(spec, strict_approved_only=strict_approved_only)
    if dry_run:
        print(code)
        return output_path
    if output_path.exists() and not overwrite:
        raise SystemExit(f"출력 파일이 이미 존재합니다: {output_path}")
    write_dag_file(output_path, code)
    return output_path


def main() -> int:
    args = build_parser().parse_args()
    input_path = Path(args.input).expanduser()
    if not input_path.exists():
        print(f"[FATAL] 입력 Excel을 찾을 수 없습니다: {input_path}", file=sys.stderr)
        return 2

    specs = load_excel(input_path, args.seq_names)
    if not specs:
        print("[FATAL] 생성할 DAG 대상이 없습니다.", file=sys.stderr)
        return 2

    if args.output and (args.split_by_seq or len(specs) > 1):
        print("[FATAL] --output은 단일 DAG 생성일 때만 사용할 수 있습니다.", file=sys.stderr)
        return 2

    output_dir = Path(args.output_dir).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)

    written: List[Path] = []
    if args.output:
        spec = next(iter(specs.values()))
        output_path = Path(args.output).expanduser()
        written.append(generate_one(spec, output_path, args.strict_approved_only, args.dry_run, args.overwrite))
    else:
        if args.split_by_seq or len(specs) > 1:
            for spec in specs.values():
                output_path = output_dir / f"{spec.dag_id}.py"
                written.append(generate_one(spec, output_path, args.strict_approved_only, args.dry_run, args.overwrite))
        else:
            spec = next(iter(specs.values()))
            output_path = output_dir / f"{spec.dag_id}.py"
            written.append(generate_one(spec, output_path, args.strict_approved_only, args.dry_run, args.overwrite))

    if not args.dry_run:
        for path in written:
            print(f"[OK] {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

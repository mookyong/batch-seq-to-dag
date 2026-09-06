#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
두 테이블(`seq_operator_questionnaire`, `seq_operator_questionnaire_tag`)을 읽어
DAG 자동 생성용 Excel 워크북(.xlsx)을 만든다.

출력 시트:
  - DAG_Mapping
  - Sensor_Notify
  - Validation_Policy
  - Review_Log

기본 동작:
  - DB의 모든 질문지 행을 읽는다.
  - 각 행에 대해 자동 추론값과 사람이 채워야 할 칸을 함께 쓴다.
  - 하나의 워크북에 시트별 표 형태로 저장한다.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import pymysql
except ImportError as exc:  # pragma: no cover
    raise SystemExit("PyMySQL이 필요합니다: pip install PyMySQL") from exc

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    load_dotenv = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl이 필요합니다: pip install openpyxl") from exc


SHEETS: Dict[str, List[str]] = {
    "DAG_Mapping": [
        "seq_name",
        "business_name",
        "dag_id",
        "dag_name",
        "schedule_type",
        "schedule_interval",
        "priority",
        "dbt_group",
        "dbt_selector",
        "model_mapping",
    ],
    "Sensor_Notify": [
        "seq_name",
        "start_mode",
        "sensor_type",
        "sensor_target",
        "external_dependency",
        "notify_on_success",
        "notify_on_failure",
        "notify_channel",
        "notify_recipients",
        "failure_policy",
    ],
    "Validation_Policy": [
        "seq_name",
        "validation_mode",
        "validation_timing",
        "comparison_reference",
        "retry_policy",
        "exception_rule",
        "verification_comment",
        "review_status",
    ],
    "Review_Log": [
        "seq_name",
        "review_status",
        "review_comment",
        "owner",
        "reviewed_at",
    ],
}

COL_STATUS: Dict[str, Dict[str, str]] = {
    "DAG_Mapping": {
        "seq_name": "auto",
        "business_name": "auto",
        "dag_id": "auto",
        "dag_name": "auto",
        "schedule_type": "suggest",
        "schedule_interval": "manual",
        "priority": "suggest",
        "dbt_group": "suggest",
        "dbt_selector": "suggest",
        "model_mapping": "manual",
    },
    "Sensor_Notify": {
        "seq_name": "auto",
        "start_mode": "suggest",
        "sensor_type": "manual",
        "sensor_target": "manual",
        "external_dependency": "suggest",
        "notify_on_success": "manual",
        "notify_on_failure": "manual",
        "notify_channel": "manual",
        "notify_recipients": "manual",
        "failure_policy": "suggest",
    },
    "Validation_Policy": {
        "seq_name": "auto",
        "validation_mode": "suggest",
        "validation_timing": "manual",
        "comparison_reference": "suggest",
        "retry_policy": "suggest",
        "exception_rule": "manual",
        "verification_comment": "manual",
        "review_status": "manual",
    },
    "Review_Log": {
        "seq_name": "auto",
        "review_status": "manual",
        "review_comment": "manual",
        "owner": "auto",
        "reviewed_at": "auto",
    },
}

FILL = {
    "auto": PatternFill("solid", fgColor="D9EAD3"),
    "suggest": PatternFill("solid", fgColor="FFF2CC"),
    "manual": PatternFill("solid", fgColor="FCE5CD"),
    "review": PatternFill("solid", fgColor="F4CCCC"),
    "header": PatternFill("solid", fgColor="D9E1F2"),
}

HEADER_FONT = Font(bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(vertical="top", wrap_text=True)

AMBIGUOUS_MARKERS = ("모름/확인 필요", "상황별 판단", "확인 필요", "일부 가능")


def normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\xa0", " ").strip())


def slugify(value: str) -> str:
    value = normalize_whitespace(value).lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value


def safe_filename(value: str) -> str:
    value = normalize_whitespace(value)
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", "_", value).strip("._")
    return value or "output"


def load_environment(env_file: str, required: bool) -> None:
    path = Path(env_file).expanduser() if env_file else None
    if path and path.exists():
        if load_dotenv is None:
            raise SystemExit("python-dotenv가 필요합니다: pip install python-dotenv")
        load_dotenv(dotenv_path=path, override=False)
    elif required:
        needed = ("MARIADB_USER", "MARIADB_PASSWORD", "MARIADB_DATABASE")
        if not all(os.getenv(k, "").strip() for k in needed):
            raise SystemExit(
                f".env 파일을 찾을 수 없습니다: {path}. 또는 MARIADB_USER/MARIADB_PASSWORD/MARIADB_DATABASE를 환경변수로 주입하세요."
            )


def env_required(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise SystemExit(f"필수 환경변수가 없습니다: {name}")
    return value


def connect_db():
    host = os.getenv("MARIADB_HOST", "127.0.0.1").strip() or "127.0.0.1"
    port = int(os.getenv("MARIADB_PORT", "3307"))
    user = env_required("MARIADB_USER")
    password = env_required("MARIADB_PASSWORD")
    database = env_required("MARIADB_DATABASE")
    connect_timeout = int(os.getenv("MARIADB_CONNECT_TIMEOUT", "10"))

    return pymysql.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        charset="utf8mb4",
        autocommit=False,
        connect_timeout=connect_timeout,
        cursorclass=pymysql.cursors.DictCursor,
    )


def load_json_list(value: Any) -> List[str]:
    if not value:
        return []
    if isinstance(value, list):
        return [str(v) for v in value]
    try:
        data = json.loads(value)
    except Exception:
        return []
    if isinstance(data, list):
        return [str(v) for v in data]
    return []


def fetch_questionnaires(conn: Any, seq_names: Optional[Sequence[str]]) -> List[Dict[str, Any]]:
    sql = "SELECT * FROM seq_operator_questionnaire"
    params: List[Any] = []
    if seq_names:
        placeholders = ", ".join(["%s"] * len(seq_names))
        sql += f" WHERE seq_name IN ({placeholders})"
        params.extend(seq_names)
    sql += " ORDER BY seq_name"
    with conn.cursor() as cur:
        cur.execute(sql, params)
        return list(cur.fetchall())


def fetch_tags(conn: Any, questionnaire_ids: Sequence[int]) -> Dict[int, List[Dict[str, Any]]]:
    if not questionnaire_ids:
        return {}
    placeholders = ", ".join(["%s"] * len(questionnaire_ids))
    sql = (
        "SELECT questionnaire_id, tag_category, tag_name, source_field, source_value, rule_code, confidence, requires_review "
        "FROM seq_operator_questionnaire_tag "
        f"WHERE questionnaire_id IN ({placeholders}) "
        "ORDER BY questionnaire_id, tag_name"
    )
    with conn.cursor() as cur:
        cur.execute(sql, list(questionnaire_ids))
        tags: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
        for row in cur.fetchall():
            tags[int(row["questionnaire_id"])].append(row)
        return tags


def tag_names(tags: Sequence[Dict[str, Any]]) -> set[str]:
    return {str(t.get("tag_name", "")) for t in tags if t.get("tag_name")}


def infer_schedule_type(q: Dict[str, Any], names: set[str]) -> str:
    if "cycle_daily" in names:
        return "daily"
    if "cycle_weekly" in names:
        return "weekly"
    if "cycle_monthly" in names:
        return "monthly"
    if "cycle_adhoc" in names or "run_manual" in names:
        return "manual"
    if {"trigger_file", "trigger_dependency", "trigger_data"} & names:
        return "event"
    return ""


def infer_priority(q: Dict[str, Any], names: set[str]) -> str:
    raw = normalize_whitespace(str(q.get("operational_importance_raw", "")))
    code = normalize_whitespace(str(q.get("operational_importance_code", "")))
    if "critical" in names or "상" in raw or code == "상":
        return "high"
    if "중" in raw or code == "중":
        return "medium"
    if "하" in raw or code == "하":
        return "low"
    return "medium"


def infer_dbt_group(seq_name: str) -> str:
    base = re.sub(r"^seq[_-]?", "", seq_name, flags=re.IGNORECASE)
    return slugify(base)


def infer_start_mode(q: Dict[str, Any], names: set[str]) -> str:
    if "run_manual" in names:
        return "manual"
    if "trigger_file" in names:
        return "file_sensor"
    if "trigger_dependency" in names:
        return "dependency_sensor"
    if "trigger_schedule" in names or any(x in names for x in ("cycle_daily", "cycle_weekly", "cycle_monthly")):
        return "time_based"
    if "trigger_data" in names:
        return "event"
    return ""


def infer_sensor_type(start_mode: str) -> str:
    return {
        "file_sensor": "FileSensor",
        "dependency_sensor": "ExternalTaskSensor",
        "event": "CustomSensor",
    }.get(start_mode, "")


def infer_sensor_target(q: Dict[str, Any], start_mode: str) -> str:
    if start_mode == "file_sensor":
        return normalize_whitespace(str(q.get("file_data_wait", "") or q.get("external_dependency", "")))
    if start_mode == "dependency_sensor":
        return normalize_whitespace(str(q.get("external_dependency", "")))
    if start_mode == "event":
        return normalize_whitespace(str(q.get("external_dependency", "") or q.get("file_data_wait", "")))
    return ""


def infer_failure_policy(q: Dict[str, Any], names: set[str]) -> str:
    raw = normalize_whitespace(str(q.get("job_failure_action_raw", "")))
    if "SEQ 전체 종료" in raw:
        return "stop"
    if "다음 단계 계속" in raw:
        return "continue"
    if "운영자 확인 후 처리" in raw:
        return "manual_review"
    if "critical" in names:
        return "stop"
    return ""


def infer_validation_mode(q: Dict[str, Any], names: set[str]) -> str:
    methods = load_json_list(q.get("verification_method_json"))
    mapping = {
        "건수": "count",
        "금액/수량 합계": "sum",
        "주요 Key/샘플": "sample",
        "보고서/화면": "report",
        "파일": "file",
        "기타": "manual",
    }
    mapped = [mapping[m] for m in methods if m in mapping]
    if len(mapped) == 1:
        return mapped[0]
    if mapped:
        return "dbt_test"
    if any(q.get(k) for k in ("verification_timing", "main_verification_target", "null_duplicate_missing_check")):
        return "dbt_test"
    return ""


def infer_comparison_reference(q: Dict[str, Any]) -> str:
    refs = load_json_list(q.get("comparison_reference_json"))
    mapping = {
        "원천 데이터": "source",
        "기존 DW/DM": "existing_dw",
        "전일/전월 결과": "prior_period",
        "보고서": "report",
        "외부 파일/시스템": "external",
        "기타": "other",
    }
    mapped = [mapping[r] for r in refs if r in mapping]
    if mapped:
        return ", ".join(mapped)
    return normalize_whitespace(str(q.get("comparison_reference_raw", "")))


def infer_retry_policy(q: Dict[str, Any]) -> str:
    return normalize_whitespace(str(q.get("retry_criteria", "")))


def infer_exception_rule(q: Dict[str, Any]) -> str:
    return normalize_whitespace(str(q.get("exception_execution_rule", "")))


def infer_verification_comment(q: Dict[str, Any]) -> str:
    parts = [
        normalize_whitespace(str(q.get("main_verification_target", ""))),
        normalize_whitespace(str(q.get("null_duplicate_missing_check", ""))),
        normalize_whitespace(str(q.get("verification_failure_action", ""))),
    ]
    parts = [p for p in parts if p]
    return " / ".join(parts)


def review_comment(q: Dict[str, Any], dag_row: Dict[str, Any], sensor_row: Dict[str, Any], validation_row: Dict[str, Any]) -> str:
    missing: List[str] = []
    for key in ("schedule_interval", "sensor_type", "sensor_target", "model_mapping", "validation_timing", "notify_channel", "notify_recipients", "exception_rule"):
        if not normalize_whitespace(str(dag_row.get(key, "") or sensor_row.get(key, "") or validation_row.get(key, ""))):
            missing.append(key)

    ambiguities: List[str] = []
    for field in (
        "reprocessing_detail",
        "exception_execution_rule",
        "interview_additional_check",
        "must_keep_after_transition",
        "history_comparable_period",
    ):
        value = normalize_whitespace(str(q.get(field, "")))
        if any(marker in value for marker in AMBIGUOUS_MARKERS):
            ambiguities.append(field)

    notes: List[str] = []
    if missing:
        notes.append("확인 필요: " + ", ".join(sorted(set(missing))))
    if ambiguities:
        notes.append("모호성: " + ", ".join(sorted(set(ambiguities))))
    if not notes:
        notes.append("자동 초안 생성됨")
    return " ; ".join(notes)


def build_rows(questionnaire: Dict[str, Any], tags: Sequence[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    names = tag_names(tags)
    seq_name = normalize_whitespace(str(questionnaire.get("seq_name", "")))
    business_name = normalize_whitespace(str(questionnaire.get("business_name", "")))
    dag_id = slugify(seq_name)
    dag_name = f"{seq_name} - {business_name}" if business_name else seq_name
    schedule_type = infer_schedule_type(questionnaire, names)
    start_mode = infer_start_mode(questionnaire, names)
    sensor_type = infer_sensor_type(start_mode)
    sensor_target = infer_sensor_target(questionnaire, start_mode)
    priority = infer_priority(questionnaire, names)
    dbt_group = infer_dbt_group(seq_name)
    dbt_selector = f"tag:{dbt_group}" if dbt_group else ""
    failure_policy = infer_failure_policy(questionnaire, names)
    validation_mode = infer_validation_mode(questionnaire, names)
    comparison_reference = infer_comparison_reference(questionnaire)
    retry_policy = infer_retry_policy(questionnaire)
    exception_rule = infer_exception_rule(questionnaire)
    verification_comment_text = infer_verification_comment(questionnaire)

    dag_row = {
        "seq_name": seq_name,
        "business_name": business_name,
        "dag_id": dag_id,
        "dag_name": dag_name,
        "schedule_type": schedule_type,
        "schedule_interval": "",
        "priority": priority,
        "dbt_group": dbt_group,
        "dbt_selector": dbt_selector,
        "model_mapping": "",
    }

    sensor_row = {
        "seq_name": seq_name,
        "start_mode": start_mode,
        "sensor_type": sensor_type,
        "sensor_target": sensor_target,
        "external_dependency": normalize_whitespace(str(questionnaire.get("external_dependency", ""))),
        "notify_on_success": "Y" if normalize_whitespace(str(questionnaire.get("success_failure_notification", ""))) else "",
        "notify_on_failure": "Y" if normalize_whitespace(str(questionnaire.get("success_failure_notification", "")) or normalize_whitespace(str(questionnaire.get("job_failure_action_raw", "")))) else "",
        "notify_channel": "",
        "notify_recipients": "",
        "failure_policy": failure_policy,
    }

    validation_row = {
        "seq_name": seq_name,
        "validation_mode": validation_mode,
        "validation_timing": normalize_whitespace(str(questionnaire.get("verification_timing", ""))),
        "comparison_reference": comparison_reference,
        "retry_policy": retry_policy,
        "exception_rule": exception_rule,
        "verification_comment": verification_comment_text,
        "review_status": "needs_review",
    }

    review_row = {
        "seq_name": seq_name,
        "review_status": "needs_review",
        "review_comment": review_comment(questionnaire, dag_row, sensor_row, validation_row),
        "owner": normalize_whitespace(str(questionnaire.get("owner", ""))),
        "reviewed_at": str(questionnaire.get("imported_at", ""))[:19].replace("T", " "),
    }

    return {
        "DAG_Mapping": dag_row,
        "Sensor_Notify": sensor_row,
        "Validation_Policy": validation_row,
        "Review_Log": review_row,
    }


def apply_sheet_style(ws, sheet_name: str, columns: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = FILL["header"]
        cell.border = BORDER
        cell.alignment = CENTER

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name, "")
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = CENTER
            status = COL_STATUS[sheet_name].get(col_name, "manual")
            cell.fill = FILL.get(status, FILL["manual"])
            if sheet_name == "Review_Log" and col_name == "review_status" and str(value).strip() == "needs_review":
                cell.fill = FILL["review"]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row in rows:
            value = row.get(col_name, "")
            if value is None:
                text = ""
            else:
                text = str(value)
            for line in text.splitlines() or [""]:
                max_len = max(max_len, len(line))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)


def add_validations(ws, sheet_name: str, max_row: int) -> None:
    validations = {
        "DAG_Mapping": {
            "schedule_type": ["daily", "weekly", "monthly", "manual", "event"],
        },
        "Sensor_Notify": {
            "start_mode": ["file_sensor", "dependency_sensor", "manual", "time_based", "event"],
            "sensor_type": ["FileSensor", "ExternalTaskSensor", "CustomSensor"],
            "notify_on_success": ["Y", "N"],
            "notify_on_failure": ["Y", "N"],
            "notify_channel": ["Slack", "Email", "Callback", "Teams"],
            "failure_policy": ["stop", "continue", "manual_review"],
        },
        "Validation_Policy": {
            "validation_mode": ["count", "sum", "sample", "report", "file", "dbt_test", "manual"],
            "review_status": ["approved", "needs_review", "blocked"],
        },
        "Review_Log": {
            "review_status": ["approved", "needs_review", "blocked"],
        },
    }

    if sheet_name not in validations:
        return

    header = {name: idx + 1 for idx, name in enumerate(SHEETS[sheet_name])}
    for col_name, values in validations[sheet_name].items():
        col_idx = header[col_name]
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        dv.prompt = f"{col_name} 값을 선택하세요"
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{max_row}")


def write_workbook(output_path: Path, questionnaires: List[Dict[str, Any]], tags_by_qid: Dict[int, List[Dict[str, Any]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    wb.properties.title = "DAG Generation Input"
    wb.properties.subject = "SEQ DAG 자동 생성용 Excel"
    wb.properties.creator = "OpenCode"

    rows_by_sheet: Dict[str, List[Dict[str, Any]]] = {sheet: [] for sheet in SHEETS}
    for q in questionnaires:
        qid = int(q["id"])
        built = build_rows(q, tags_by_qid.get(qid, []))
        for sheet_name, row in built.items():
            rows_by_sheet[sheet_name].append(row)

    for sheet_name, columns in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        rows = rows_by_sheet[sheet_name]
        apply_sheet_style(ws, sheet_name, columns, rows)
        add_validations(ws, sheet_name, max(2, len(rows) + 1))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_split_workbooks(output_dir: Path, questionnaires: List[Dict[str, Any]], tags_by_qid: Dict[int, List[Dict[str, Any]]]) -> List[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    written: List[Path] = []
    for q in questionnaires:
        qid = int(q["id"])
        seq_name = normalize_whitespace(str(q.get("seq_name", "")))
        file_name = f"{safe_filename(seq_name)}.xlsx"
        out_path = output_dir / file_name
        write_workbook(out_path, [q], {qid: tags_by_qid.get(qid, [])})
        written.append(out_path)
    return written


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="두 테이블을 기반으로 DAG 생성용 Excel(.xlsx) 워크북을 만든다")
    p.add_argument("--output", default="dag_generation_input.xlsx", help="출력 .xlsx 파일 경로")
    p.add_argument("--split-by-seq", action="store_true", help="seq_name별로 개별 .xlsx 파일을 생성한다")
    p.add_argument("--output-dir", default="dag_generation_input_out", help="--split-by-seq 사용 시 출력 디렉터리")
    p.add_argument("--seq-name", action="append", dest="seq_names", help="대상 seq_name (여러 번 지정 가능)")
    p.add_argument("--env-file", default=os.getenv("ENV_FILE", ".env"), help="MariaDB 접속 환경변수 파일 경로")
    p.add_argument("--overwrite", action="store_true", help="출력 파일이 있어도 덮어쓴다")
    return p


def main() -> int:
    args = build_parser().parse_args()
    load_environment(args.env_file, required=True)

    output_path = Path(args.output).expanduser()
    output_dir = Path(args.output_dir).expanduser()
    if args.split_by_seq:
        if output_dir.exists() and any(output_dir.iterdir()) and not args.overwrite:
            print(f"[FATAL] 출력 디렉터리가 이미 비어 있지 않습니다: {output_dir}", file=sys.stderr)
            return 2
    else:
        if output_path.exists() and not args.overwrite:
            print(f"[FATAL] 출력 파일이 이미 존재합니다: {output_path}", file=sys.stderr)
            return 2

    conn = connect_db()
    try:
        questionnaires = fetch_questionnaires(conn, args.seq_names)
        if not questionnaires:
            print("처리할 질문지 데이터가 없습니다.", file=sys.stderr)
            return 2
        qids = [int(q["id"]) for q in questionnaires]
        tags_by_qid = fetch_tags(conn, qids)
        if args.split_by_seq:
            written = write_split_workbooks(output_dir, questionnaires, tags_by_qid)
            print(f"[OK] {output_dir} 아래 {len(written)}개 파일 생성 완료")
            for path in written:
                print(f"     {path}")
        else:
            write_workbook(output_path, questionnaires, tags_by_qid)
            print(f"[OK] {output_path} 생성 완료")
        print(f"     questionnaires={len(questionnaires)}, sheets={len(SHEETS)}")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
